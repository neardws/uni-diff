import os
from typing import List
from .base import BaseConverter, ConvertedDocument, TextBlock


class XLSXConverter(BaseConverter):
    """Converter for Microsoft Excel XLSX files."""

    @property
    def supported_extensions(self) -> List[str]:
        return ['.xlsx', '.xls']

    def convert(self, file_path: str) -> ConvertedDocument:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        blocks = []
        full_text = ""

        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True)

            page_num = 0
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                full_text += f"=== Sheet: {sheet_name} ===\n"
                blocks.append(TextBlock(
                    text=f"=== Sheet: {sheet_name} ===",
                    page=page_num,
                    x=0,
                    y=0,
                    width=200,
                    height=16,
                    metadata={'type': 'sheet_header', 'sheet': sheet_name}
                ))

                y_offset = 20
                for row_idx, row in enumerate(sheet.iter_rows(), start=1):
                    row_values = []
                    for cell in row:
                        value = cell.value
                        if value is not None:
                            row_values.append(str(value))
                        else:
                            row_values.append("")

                    if any(v.strip() for v in row_values):
                        text = " | ".join(row_values)
                        full_text += text + "\n"
                        blocks.append(TextBlock(
                            text=text,
                            page=page_num,
                            x=0,
                            y=y_offset,
                            width=len(text) * 7,
                            height=14,
                            metadata={
                                'type': 'row',
                                'sheet': sheet_name,
                                'row': row_idx
                            }
                        ))
                    y_offset += 14

                page_num += 1
                full_text += "\n"

            wb.close()

        except ImportError:
            try:
                import csv
                import subprocess
                result = subprocess.run(
                    ['ssconvert', '--export-type=Gnumeric_stf:stf_csv', file_path, 'fd://1'],
                    capture_output=True,
                    text=True
                )
                if result.returncode == 0:
                    full_text = result.stdout
                    lines = full_text.split('\n')
                    y_offset = 0
                    for line in lines:
                        if line.strip():
                            blocks.append(TextBlock(
                                text=line,
                                page=0,
                                x=0,
                                y=y_offset,
                                width=len(line) * 7,
                                height=12
                            ))
                        y_offset += 12
                else:
                    raise RuntimeError("ssconvert failed")
            except (FileNotFoundError, RuntimeError):
                raise RuntimeError(
                    "Neither openpyxl nor ssconvert is available. "
                    "Install openpyxl: pip install openpyxl"
                )

        return ConvertedDocument(
            blocks=blocks,
            full_text=full_text,
            page_count=max(1, page_num),
            source_path=file_path,
            source_type='xlsx'
        )
